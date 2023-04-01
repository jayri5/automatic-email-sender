from django.shortcuts import render
import pandas as pd
import smtplib as sm
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl

# Create your views here.
def sendmail(request):
    
    data = openpyxl.load_workbook("templates/students.xlsx")
    print(type(data))
    
    #datas = data.active
    #datas.append(['jayantabagchi77@gmail.com'])
    #data.save("templates/students.xlsx")
    
    #wb = openpyxl.Workbook()
    #ws = wb.active
    #mylist = ['dog', 'cat', 'fish', 'bird']
    #ws.append(mylist)
    #wb.save('myFile.xlsx')
    
    dataframe1 = data.active
    list_of_mails = []
    
    # Iterate the loop to read the cell values
    for row in range(0, dataframe1.max_row):
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            list_of_mails.append(col[row].value)
            print(col[row].value)

    #email_col = data.get("email")
    #print(email_col)

    #list_of_mails = list(email_col)
    print(list_of_mails)

    try:
        server = sm.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login("jayri.researcher5@gmail.com", "edgyfpzjoadtkmgh")
        from_ = "jayri.researcher5@gmail.com"
        to_ = list_of_mails
        message = MIMEMultipart("alternative")
        message['Subject'] = "ed-tech sales pitch"
        message["from"] = "jayri.researcher5@gmail.com"
        
        file = open("templates/test.txt")
        print(file.readable())
            
        #file1 = open("templates/test.txt","r")
        #with open('templates/test.txt') as f:
            #mylist = [line.rstrip('\n') for line in f]
        #html = mylist
        #print(html)
        
        f = open('templates/test.txt', 'r')
        if f.mode == 'r':
            contents =f.read()
        print(contents)
        
        
        #part2 = MIMEText(html, "html")
        part2 = MIMEText(contents)
        message.attach(part2)
        server.sendmail(from_, to_, message.as_string())
        print("done")
    except Exception as e:
        print(e)

    return render(request, 'pd.html')

def home(request):
    return render(request, 'index.html')

def pd(request):
    return render(request, 'pd.html')

def ib(request):
    name = request.POST['name']
    print(name)
    data = openpyxl.load_workbook("templates/students.xlsx")
    
    datas = data.active
    datas.append([name])
    data.save("templates/students.xlsx")
    
    return render(request, 'ib.html')